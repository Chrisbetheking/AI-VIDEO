from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from utils import parse_metric, split_tags

COLUMNS = [
    "platform",
    "account_name",
    "account_url",
    "account_id",
    "followers_count",
    "total_likes",
    "last_post_at",
    "video_title",
    "video_url",
    "published_at",
    "like_count",
    "comment_count",
    "favorite_count",
    "share_count",
    "view_count",
    "thumbnail_url",
    "is_pinned",
    "tags",
    "notes",
]


def write_excel(rows: list[dict[str, Any]], output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "accounts"
    ws.append(COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in COLUMNS])
    for col in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in col) + 2, 48)
        ws.column_dimensions[col[0].column_letter].width = width
    wb.save(path)
    return path


def read_excel_rows(path: str | Path) -> list[dict[str, Any]]:
    wb = load_workbook(path)
    ws = wb["accounts"] if "accounts" in wb.sheetnames else wb.active
    header = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {header[i]: values[i] if i < len(values) else None for i in range(len(header))}
        if not any(row.values()):
            continue
        rows.append(row)
    return rows


def excel_rows_to_accounts(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        account_url = str(row.get("account_url") or "").strip()
        account_name = str(row.get("account_name") or "").strip()
        if not account_url and not account_name:
            continue
        key = account_url or account_name
        account = grouped.setdefault(
            key,
            {
                "name": account_name or "未命名账号",
                "platform": str(row.get("platform") or "抖音").strip(),
                "url": account_url,
                "account_id": str(row.get("account_id") or "").strip(),
                "tags": split_tags(row.get("tags")),
                "notes": str(row.get("notes") or "").strip(),
                "followers_count": parse_metric(row.get("followers_count")),
                "total_likes": parse_metric(row.get("total_likes")),
                "last_post_at": str(row.get("last_post_at") or "").strip(),
                "recent_items": [],
            },
        )
        if row.get("video_url") or row.get("video_title"):
            account["recent_items"].append(
                {
                    "platform": account["platform"],
                    "account_name": account["name"],
                    "account_url": account["url"],
                    "title": str(row.get("video_title") or "").strip(),
                    "url": str(row.get("video_url") or "").strip(),
                    "published_at": str(row.get("published_at") or "").strip(),
                    "like_count": parse_metric(row.get("like_count")),
                    "comment_count": parse_metric(row.get("comment_count")),
                    "favorite_count": parse_metric(row.get("favorite_count")),
                    "share_count": parse_metric(row.get("share_count")),
                    "view_count": parse_metric(row.get("view_count")),
                    "thumbnail_url": str(row.get("thumbnail_url") or "").strip(),
                    "is_pinned": str(row.get("is_pinned") or "").strip().lower() in {"1", "true", "yes", "是", "置顶"},
                    "tags": account["tags"],
                }
            )
    return list(grouped.values())
